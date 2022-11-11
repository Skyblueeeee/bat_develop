# 仅供娱乐，切勿认真，源码辣眼睛，建议不看
import os,sys
import openpyxl
import shutil
import chinese_calendar
import datetime



####################################### 个性化更改 ###########################################
# 已两个项目号举例，可进行适当的修改
work_time = [4,4]
project = ["项目号1","项目号2"]
work_item = ["模型训练及调优","集成调试（算法及应用层测试）"]
work_item_plus = ["实际检出测试、跟踪及分析","算法及应用层测试"]
leader_plan = ["项目点检","整理数据以及模型训练"]
####################################### 个性化更改 ###########################################




####################################### 切勿更改 ###########################################
# 该月工时月份
month = str(sys.argv[1])
name = str(sys.argv[2])
state = "已完成"
work_time_S = sum(work_time)
# EXCEL地址
old_xlsx = r"src\Excel\工时计划_2022模板.xlsx"


def workdate(month):
    workday = []
    # 本月的起始时间,会自动获取工作日
    import calendar
    year = int(month[:4])
    month = int(month[4:])
    monthRange = calendar.monthrange(year,month)
    start_time = datetime.date(year,month, 1)
    end_time = datetime.date(year,month,monthRange[1])
    workdays = chinese_calendar.get_workdays(start_time,end_time)
    for day in workdays:
        workday.append(str(day).replace("-",""))
    return workday

def xls2xlsx(old_file,new_file):
    """
    将xls文件另存为xlsx文件
    
    :param file_name: 要转换的文件路径
    :returns: new_excel_file_path 返回新的xlsx文件的路径
    """
    excel_file_path = old_file
    import win32com.client
    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(excel_file_path)
 
    new_excel_file_path = new_file
    if os.path.exists(new_excel_file_path):  # 先删掉新复制的文件
        os.remove(new_excel_file_path)
    wb.SaveAs(new_excel_file_path, FileFormat=51)# 51 表示的是xlsx格式
    print(f"已新建{new_file[44:]}的工时文件")
    wb.Close()
    excel.Application.Quit()

def create_sheet(date):
    new_sheet = workbook[date]
    # 写入名字
    new_sheet.cell(row = 3,column = 2).value = name
    # 写入日期
    new_sheet.cell(row = 3,column = 5).value = int(date)
    # 写入项目号
    new_sheet.cell(row = 5,column = 1).value = project[0]
    new_sheet.cell(row = 6,column = 1).value = project[1]
    # 写入工作类别
    new_sheet.cell(row = 5,column = 2).value = work_item[0]
    new_sheet.cell(row = 6,column = 2).value = work_item[1]
    # 写入详细工作类别
    new_sheet.cell(row = 5,column = 3).value = work_item_plus[0]
    new_sheet.cell(row = 6,column = 3).value = work_item_plus[1]
    # 写入主管计划内容
    new_sheet.cell(row = 5,column = 4).value = leader_plan[0]
    new_sheet.cell(row = 6,column = 4).value = leader_plan[1]
    # 写入评定耗时
    new_sheet.cell(row = 5,column = 7).value = work_time[0]
    new_sheet.cell(row = 6,column = 7).value = work_time[1]
    # 写入完成状态
    new_sheet.cell(row = 5,column = 8).value = state
    new_sheet.cell(row = 6,column = 8).value = state
    # 写入实际耗时
    new_sheet.cell(row = 5,column = 9).value = work_time[0]
    new_sheet.cell(row = 6,column = 9).value = work_time[1]
    # 写入打卡时间
    new_sheet.cell(row = 3,column = 7).value = work_time_S

    print(f"工作簿{date}写入完成")
    workbook.save(new_xlsx)

if __name__ == "__main__":
    new_xlsx = os.path.join(name + "_工时计划"+ month + ".xlsx")
    print(new_xlsx)
    if not os.path.exists(new_xlsx):
        shutil.copy(old_xlsx,new_xlsx)

    workbook = openpyxl.load_workbook(new_xlsx)
    sheet = workbook["模板"]
    sn = workbook.sheetnames

    date_list = workdate(month)

    for date in date_list:
        if date not in sn:
            new_sheet = workbook.copy_worksheet(sheet,date) 
            print(f"已新建{date}工作簿")
            create_sheet(date)
